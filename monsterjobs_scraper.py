#! python3
# job_scraper.py - Gathers job listings

import os
import time
import urllib
import urllib.parse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys

import indeedjobs_scraper


def find_jobs_from(website, job_title, location, desired_characs, filename="results.xlsx"):
    if website == 'Monster':
        job_soup = load_monster_jobs_div(job_title, location)
        jobs_list, num_listings = extract_job_information_monster(job_soup, desired_characs)
        save_jobs_to_excel(jobs_list, filename)

# Saves Gathered Scraped info to spreadsheet
def save_jobs_to_excel(jobs_list, filename):
    links = jobs_list['links']
    wb = Workbook()
    ws = wb.active 
    ws.title = 'Monster Jobs Results'
    
    for r in dataframe_to_rows(pd.DataFrame(jobs_list), index=False, header=True):
        ws.append(r)
        
    i = 1
    for link in links:
        i = i + 1
        ws['C{}'.format(i)].value = '=HYPERLINK("{}", "{}")'.format(("%s" % link), "Link to Apply")
    
    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'
        
    wb.save(os.getcwd() + "/Desktop/" + filename)

# Gathers raw site data of needed components
def load_monster_jobs_div(job_title, location):
    job_list = []
    options = Options()
    options.headless = True
    options.add_argument("--window-size=1920, 1080")
    options.add_argument("start_maximized")
    options.add_experimental_option("prefs", {"profile.managed_default_content_settings.images": 2})
    driver = webdriver.Chrome(options=options)
    getVars = {'q': job_title, '&where': location}
    for i in range(1,5):
        url = ('https://www.monster.com/jobs/search?' + urllib.parse.urlencode(getVars) + '&page=%s&so=m.h.s' % i)
        driver.get(url)
        time.sleep(2)
        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")
        job_soup = soup.find(id="JobCardGrid")
        job_list.append(job_soup)
    return job_list
    
# Extracts Gathered info from site and prepares it for saving to spreadsheet
def extract_job_information_monster(job_soup, desired_characs):
    cols = []
    extracted_info = []
    titles = []
    companies = []
    links = []
    dates = []
    for soup in job_soup:
        job_elems = soup.find_all('article', class_= "job-cardstyle__JobCardComponent-sc-1mbmxes-0")
        if 'titles' in desired_characs:
            cols.append('titles')
            for job_elem in job_elems:
                titles.append(extract_job_titles_monster(job_elem))
            extracted_info.append(titles)
        
        if 'companies' in desired_characs:
            cols.append('companies')
            for job_elem in job_elems:
                companies.append(extract_company_monster(job_elem))
            extracted_info.append(companies)
            
        if 'links' in desired_characs:
            cols.append('links')
            for job_elem in job_elems:
                links.append(extract_link_monster(job_elem))
            extracted_info.append(links)
            
        if 'date_listed' in desired_characs:
            cols.append('date_listed')
            for job_elem in job_elems:
                dates.append(extract_date_monster(job_elem))
            extracted_info.append(dates)
                
    jobs_list = {}
    
    for j in range(len(cols)):
        jobs_list[cols[j]] = extracted_info[j]
    
    num_listings = len(extracted_info[0])
    return jobs_list, num_listings
            
def extract_job_titles_monster(job_elem):
    title_elem = job_elem.find('a', class_='job-cardstyle__JobCardTitle-sc-1mbmxes-2')
    title = title_elem.text.strip()
    return title
    
def extract_company_monster(job_elem):
    company_elem = job_elem.find('h3', class_='job-cardstyle__JobCardCompany-sc-1mbmxes-3')
    company = company_elem.text.strip()
    return company

def extract_link_monster(job_elem):
    link = job_elem.find('a', class_='job-cardstyle__JobCardTitle-sc-1mbmxes-2')['href']
    link = 'http:' + link
    return link

def extract_date_monster(job_elem):
    date_elem = job_elem.find('span', class_='job-cardstyle__JobCardDate-sc-1mbmxes-6')
    date = date_elem.text.strip()
    return date

def main(job_type, job_location):
    print("========================================\n\nThis script will gather the job titles, links, & date posted.\n\nOnce finished, it will save the results to an excel spreadsheet on your *Desktop*.\n\n========================================")
    desired_characs = ['titles', 'companies', 'links', 'date_listed']
    find_jobs_from('Monster', job_type, job_location, desired_characs)
    indeedjobs_scraper.main(job_type, job_location)
    
    

    