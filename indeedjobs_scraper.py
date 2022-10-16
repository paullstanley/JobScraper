#! python3
# indeedjobs_scraper.py - Gathers job listings

import os
import urllib
import urllib.parse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def find_jobs_from(website, job_title, location, desired_characs, filename="results.xlsx"):
    if website == 'Indeed':
        job_soup = load_indeed_jobs_div(job_title, location)
        jobs_list, num_listings = extract_job_information_indeed(job_soup, desired_characs)
        save_jobs_to_excel(jobs_list, filename)

# Saves Gathered Scraped info to spreadsheet
def save_jobs_to_excel(jobs_list, filename):
    links = jobs_list['links']
    wb = load_workbook(os.getcwd() + "/Desktop/" + filename)
    ws = wb.create_sheet("Indeed Job Results")

    for r in dataframe_to_rows(pd.DataFrame(jobs_list), index=False, header=True):
        ws.append(r)
        
    i = 1
    for link in links:
        i = i + 1
        ws['C{}'.format(i)].value = '=HYPERLINK("{}", "{}")'.format(("%s" % link), "Link to Apply")
    
    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'
        
    wb.save(os.getcwd() + "/Desktop/" + filename)
    
def load_indeed_jobs_div(job_title, location):
    job_list = []
    getVars = {'q': job_title, '&l': location, '&fromage': 'last', '&sort': 'date'}
    for i in range(0,5):
        url = ('http://www.indeed.com/jobs?' + urllib.parse.urlencode(getVars) + '&start=%s0' % i)
        hdr = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36'}
        page = requests.get(url,headers = hdr)
        soup = BeautifulSoup(page.content, "html.parser")
        job_soup = soup.find(id="mosaic-provider-jobcards")
        job_list.append(job_soup)
    return job_list
    
def extract_job_information_indeed(job_soup, desired_characs):
    cols = []
    extracted_info = []
    titles = []
    companies = []
    links = []
    dates = []
    for soup in job_soup:
        job_elems = soup.find_all('div', class_='job_seen_beacon')
        if 'titles' in desired_characs:
            cols.append('titles')
            for job_elem in job_elems:
                titles.append(extract_job_title_indeed(job_elem))
            extracted_info.append(titles)                    
    
        if 'companies' in desired_characs:
            cols.append('companies')
            for job_elem in job_elems:
                companies.append(extract_company_indeed(job_elem))
            extracted_info.append(companies)
    
        if 'links' in desired_characs:
            cols.append('links')
            for job_elem in job_elems:
                links.append(extract_link_indeed(job_elem))
            extracted_info.append(links)
    
        if 'date_listed' in desired_characs:
            cols.append('date_listed')
            for job_elem in job_elems:
                dates.append(extract_date_indeed(job_elem))
            extracted_info.append(dates)
    
    jobs_list = {}
    
    for j in range(len(cols)):
        jobs_list[cols[j]] = extracted_info[j]
    
    num_listings = len(extracted_info[0])
    
    return jobs_list, num_listings

def extract_job_title_indeed(job_elem):
    title_elem = job_elem.find('h2', class_='jobTitle')
    title = title_elem.text.strip()
    return title

def extract_company_indeed(job_elem):
    company_elem = job_elem.find('span', class_='companyName')
    company = company_elem.text.strip()
    return company

def extract_link_indeed(job_elem):
    link = job_elem.find('a')['href']
    link = 'www.indeed.com' + link
    return link

def extract_date_indeed(job_elem):
    date_elem = job_elem.find('span', class_='date')
    date = date_elem.text.strip()
    return date

def main(job_type, job_location):
    desired_characs = ['titles', 'companies', 'links', 'date_listed']
    find_jobs_from('Indeed', job_type, job_location, desired_characs)
      